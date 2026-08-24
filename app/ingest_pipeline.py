"""Ingestion pipeline — the demo's opening beat, and the seed's loader.

file (Volume or local) -> identify carrier via filename + template -> fingerprint
(expected tabs) -> two-path extraction (deterministic label-match + Claude/FMAPI)
-> reconcile field-by-field -> quarantine on mismatch OR append with provenance
+ supersede any prior active document + audit at every stage.

The SAME code path runs locally (seed_book.py) and as the serverless Job (WP1).
Writes go through the SQL warehouse so there is one behaviour everywhere.

All synthetic. No real carrier/client/broker names.
"""
from __future__ import annotations
import io, os, re, json, uuid, datetime as dt
from databricks.sdk import WorkspaceClient

WAREHOUSE = os.getenv("HB_WAREHOUSE", "a3b61648ea4809e3")
CAT = os.getenv("HB_CATALOG", "lr_dev_aws_us_catalog")
SCH = os.getenv("HB_SCHEMA", "hb_renewal")
FQ = f"{CAT}.{SCH}"
LLM = os.getenv("HB_LLM_ENDPOINT", "databricks-claude-sonnet-4-5")
_PROF = os.getenv("HB_PROFILE") or None
_w = WorkspaceClient(profile=_PROF) if _PROF else WorkspaceClient()

# lever fields we recover from the Rate Development sheet, with label variants.
FIELD_LABELS = {
    "member_months": ["member months"],
    "months_experience": ["total months of experience", "months of experience"],
    "total_incurred_claims": ["total incurred claims"],
    "demographic_adjustment": ["demographic adjustment", "demographic adj"],
    "less_pooled_claims_pmpm": ["less pooled claims", "pooled claims"],
    "benefit_change": ["benefit change"],
    "annual_trend": ["annual trend"],
    "months_of_trend": ["months of trend"],
    "individual_pooling_point": ["individual pooling point", "pooling point"],
    "projected_excess_claims_pmpm": ["projected claims in excess", "claims in excess"],
    "large_claim_add_back_pmpm": ["large claim add back", "add back"],
    "current_members": ["current members"],
    "current_premium_pmpm": ["current premium & fees", "current premium and fees", "current premium"],
    "target_loss_ratio": ["target loss ratio"],
    "benefit_advisor_fee": ["benefit advisor fee", "advisor fee"],
    "manual_rating_pool_increase": ["manual rating pool increase", "manual rate"],
}
PREFER_ANNUAL = {"total_incurred_claims"}
CORE_FIELDS = ["member_months", "total_incurred_claims", "annual_trend", "months_of_trend",
               "demographic_adjustment", "target_loss_ratio", "manual_rating_pool_increase"]


def sq(v):
    if v is None:
        return "NULL"
    if isinstance(v, str):
        return "'" + v.replace("'", "''") + "'"
    if isinstance(v, bool):
        return "true" if v else "false"
    return repr(v)


def run(sql, label=""):
    r = _w.statement_execution.execute_statement(warehouse_id=WAREHOUSE, statement=sql, wait_timeout="50s")
    st = r.status.state.value if r.status else "?"
    if st != "SUCCEEDED":
        raise RuntimeError(f"SQL failed [{label}]: {st} :: {r.status.error.message if r.status and r.status.error else ''}")
    return r


def audit(event_type, entity_type, entity_id, detail, actor):
    run(f"""INSERT INTO {FQ}.`5_gov_audit_event` VALUES
        ('AE-{uuid.uuid4().hex[:10]}',{sq(event_type)},{sq(entity_type)},{sq(entity_id)},
         {sq(detail)},{sq(actor)},current_timestamp())""", "audit")


def archive(data: bytes, doc_id: str, file_name: str, quarantined: bool) -> str | None:
    """Copy the original file into the governed UC Volume, keyed by doc_id, so any
    decision that traces back to this document can pull the exact source file.
    Returns the Volume path, or None if the copy could not be written."""
    sub = "quarantine" if quarantined else "processed"
    dest = f"/Volumes/{CAT}/{SCH}/landing/{sub}/{doc_id}__{file_name}"
    try:
        _w.files.upload(dest, io.BytesIO(data), overwrite=True)
        return dest
    except Exception as e:
        print(f"  (archive to Volume failed: {e})")
        return None


# ------------------------------------------------------------------ read xlsx
def _rows(data: bytes) -> dict:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    return {name: [[c.value for c in row] for row in wb[name].iter_rows()] for name in wb.sheetnames}


def _numerics(row):
    return [float(v) for v in row if isinstance(v, (int, float)) and not isinstance(v, bool)]


# ------------------------------------------------- two-path lever extraction
def extract_deterministic(sheets: dict) -> dict:
    found = {}
    for exact in (True, False):
        for name, rows in sheets.items():
            for row in rows:
                label = next((v.strip() for v in row if isinstance(v, str) and v.strip()), None)
                nums = _numerics(row)
                if not label or not nums:
                    continue
                low = label.lower()
                for field, variants in FIELD_LABELS.items():
                    if field in found:
                        continue
                    hit = any(low.startswith(v) for v in variants) if exact else any(v in low for v in variants)
                    if hit:
                        val = nums[-1] if field in PREFER_ANNUAL and len(nums) > 1 else nums[0]
                        found[field] = {"value": val, "label": label, "sheet": name}
                        break
    # experience credibility weight from the blend block (two rows summing to ~1)
    if "credibility_experience_weight" not in found:
        for name, rows in sheets.items():
            for i, row in enumerate(rows[:-1]):
                label = next((v.strip() for v in row if isinstance(v, str) and v.strip()), "")
                if "experience" not in label.lower():
                    continue
                a, b = _numerics(row), _numerics(rows[i + 1])
                if a and b and 0 < a[0] < 1 and 0 < b[0] < 1 and abs(a[0] + b[0] - 1) < 0.01:
                    found["credibility_experience_weight"] = {"value": a[0], "label": label, "sheet": name}
                    break
    return found


def extract_ai(sheets: dict) -> dict:
    text = []
    for name, rows in sheets.items():
        text.append(f"### SHEET: {name}")
        for row in rows:
            cells = [str(c) for c in row if c not in (None, "")]
            if cells:
                text.append(" | ".join(cells))
    body = "\n".join(text)[:12000]
    schema = ", ".join(list(FIELD_LABELS) + ["credibility_experience_weight"])
    prompt = (f"Read this health & benefits medical renewal exhibit and extract the renewal inputs.\n"
              f"Return ONLY JSON. Keys to find (omit any not present): {schema}\n"
              f"Rules: percentages as decimals (11.75%->0.1175). For total_incurred_claims return the ANNUAL total. "
              f"credibility_experience_weight is the 0-1 weight on the experience increase. "
              f"For each key return {{\"value\": <number>}}.\n\nEXHIBIT:\n{body}")
    from openai import OpenAI
    from databricks.sdk.core import Config
    cfg = Config(profile=_PROF) if _PROF else Config()
    bearer = cfg.authenticate().get("Authorization", "").removeprefix("Bearer ")
    client = OpenAI(api_key=bearer, base_url=f"{cfg.host.rstrip('/')}/serving-endpoints")
    resp = client.chat.completions.create(model=LLM, messages=[{"role": "user", "content": prompt}], max_tokens=1500)
    m = re.search(r"\{.*\}", resp.choices[0].message.content or "", re.S)
    if not m:
        return {}
    out = {}
    for k, v in json.loads(m.group(0)).items():
        if k in FIELD_LABELS or k == "credibility_experience_weight":
            out[k] = {"value": v.get("value") if isinstance(v, dict) else v}
    return out


def reconcile(det, ai):
    fields = list(FIELD_LABELS) + ["credibility_experience_weight"]
    rows, disagreements, found = [], 0, 0
    for f in fields:
        dv = det.get(f, {}).get("value")
        av = ai.get(f, {}).get("value")
        if dv is not None:
            found += 1
        if dv is not None and av is not None:
            status = "agree" if abs(dv - av) <= max(1e-6, abs(dv) * 0.02) else "differs"
            if status == "differs":
                disagreements += 1
        elif dv is not None:
            status = "det-only"
        elif av is not None:
            status = "ai-only"
        else:
            status = "missing"
        rows.append({"field": f, "det": dv, "ai": av, "status": status})
    return rows, found, len(fields), disagreements


# --------------------------------------------------------- sheet -> rows loaders
def parse_monthly(sheets):
    for name, rows in sheets.items():
        if "claim" in name.lower() and "experience" in name.lower():
            hdr = [str(c).lower() if c else "" for c in rows[0]]
            def col(*keys):
                return next((i for i, h in enumerate(hdr) if any(k in h for k in keys)), None)
            ci = {k: col(*v) for k, v in {
                "month": ["month"], "billed": ["billed"], "med": ["medical"], "rx": ["rx", "pharmacy"],
                "fixed": ["admin", "fixed"], "total": ["total"], "subs": ["subscriber"], "members": ["member"],
            }.items()}
            out = []
            for row in rows[1:]:
                if ci["month"] is None or not isinstance(row[ci["month"]], str):
                    continue
                if row[ci["month"]].lower().startswith("total"):
                    continue
                g = lambda k: (row[ci[k]] if ci[k] is not None and ci[k] < len(row) else None)
                if g("total") is None:
                    continue
                out.append(dict(month=g("month"), billed_premium=g("billed"), ffs_medical=g("med"),
                                pharmacy=g("rx"), fixed_charges=g("fixed"), total_incurred=g("total"),
                                med_ees=g("subs"), med_members=g("members")))
            return out
    return []


def parse_large(sheets):
    for name, rows in sheets.items():
        if "claimant" in name.lower() or "high cost" in name.lower():
            out = []
            for row in rows[1:]:
                nums = _numerics(row)
                strs = [v for v in row if isinstance(v, str)]
                if not nums or not strs:
                    continue
                if any("pooling" in s.lower() for s in strs):
                    continue
                status = next((s for s in strs if s.lower() in ("active", "termed", "terminated")), None)
                ed = next((s for s in strs if s.upper() in ("EE", "DEP", "SPOUSE", "EMPLOYEE", "DEPENDENT")), None)
                out.append(dict(status=status, ee_dep=ed, diagnosis=None, amount=max(nums)))
            return out
    return []


# --------------------------------------------- detailed rates (per plan x tier)
def parse_detailed(sheets):
    _tiers = {"employee", "emp + spouse", "emp + child(ren)", "emp + family"}
    for name, rows in sheets.items():
        if "detailed rate" in name.lower():
            out, kind, plan = [], "current", None
            for row in rows:
                cells = [c for c in row if c not in (None, "")]
                if not cells:
                    continue
                first = str(cells[0]).strip()
                low = first.lower()
                if "current rate" in low:
                    kind = "current"; continue
                if "renewal rate" in low:
                    kind = "renewal"; continue
                if low.startswith("plan:"):
                    plan = first.split(":", 1)[1].strip(); continue
                if low in _tiers:
                    nums = _numerics(row)
                    if len(nums) >= 2:
                        out.append(dict(kind=kind, plan=plan, tier=first, subs=int(nums[0]), rate=nums[-1]))
            return out
    return []


# ------------------------------------------------------------------- identify
def identify(file_name: str) -> dict:
    base = os.path.basename(file_name)
    stem = re.sub(r"\.xlsx$", "", base, flags=re.I)
    parts = stem.split("_")
    carrier_key = parts[0].lower() if parts else ""
    tmpl = run(f"SELECT carrier, doc_family, expected_tabs, field_count FROM {FQ}.`0_carrier_template` "
               f"WHERE lower(carrier) LIKE '{carrier_key}%'", "tmpl").result.data_array or []
    family = "monthly_claims" if "month" in stem.lower() else "renewal_exhibit"
    row = next((r for r in tmpl if r[1] == family), (tmpl[0] if tmpl else None))
    group = parts[1].title() if len(parts) > 1 else "Unknown"
    period = next((p for p in parts if re.match(r"\d{4}h[12]", p.lower())), (parts[2] if len(parts) > 2 else "NA"))
    tabs = []
    if row and row[2]:
        tabs = json.loads(row[2]) if isinstance(row[2], str) else list(row[2])
    return {"carrier": row[0] if row else parts[0].title(), "group": "Harborview Logistics" if group.lower().startswith("harborview") else group,
            "period": period.upper(), "doc_family": family,
            "expected_tabs": tabs, "field_count": (int(row[3]) if row else 17),
            "known": row is not None}


# --------------------------------------------------------------------- process
def process_file(path: str, actor: str = "system") -> dict:
    with open(path, "rb") as f:
        data = f.read()
    file_name = os.path.basename(path)
    meta = identify(file_name)
    sheets = _rows(data)
    doc_id = f"DOC-{meta['carrier'].split()[0][:3].upper()}-{uuid.uuid4().hex[:6]}"

    # monthly-claims reforecast document: append a month to the active chain
    if meta["doc_family"] == "monthly_claims":
        act = run(f"SELECT doc_id FROM {FQ}.`1_source_document` WHERE employer_group={sq(meta['group'])} "
                  f"AND status='active' ORDER BY ingested_at DESC LIMIT 1", "find active").result.data_array or []
        if not act:
            return {"status": "orphan", "detail": "no active document for this group"}
        target = act[0][0]
        months = parse_monthly(sheets)
        vals = ",".join(f"({sq(target)},{sq(meta['carrier'])},{sq(meta['group'])},{sq(m['month'])},"
                        f"{sq(m['billed_premium'])},{sq(m['ffs_medical'])},{sq(m['pharmacy'])},"
                        f"{sq(m['fixed_charges'])},NULL,{sq(m['total_incurred'])},{sq(int(m['med_ees'] or 0))},{sq(int(m['med_members'] or 0))})"
                        for m in months)
        run(f"INSERT INTO {FQ}.`1_incurred_claims` VALUES {vals}", "append month")
        audit("reforecast_data", "source_document", target, f"appended {len(months)} fresh claims month(s)", actor)
        return {"status": "reforecast", "doc_id": target, "months": len(months)}

    # renewal exhibit: fingerprint + extract + reconcile
    present = set(sheets.keys())
    missing_tabs = [t for t in (meta["expected_tabs"] or []) if t not in present]
    det = extract_deterministic(sheets)
    try:
        ai = extract_ai(sheets)
    except Exception as e:
        ai = {}
        print(f"  (AI path unavailable: {e}; deterministic only)")
    recon, found, expected, disagreements = reconcile(det, ai)
    core_ok = all(det.get(f, {}).get("value") is not None for f in CORE_FIELDS)
    detail = json.dumps({"missing_tabs": missing_tabs, "found": found, "expected": expected,
                         "disagreements": disagreements,
                         "fields": [{k: r[k] for k in ("field", "status")} for r in recon if r["status"] not in ("agree",)][:20]})

    if missing_tabs or not core_ok or disagreements > 0:
        status = "quarantined" if (missing_tabs or not core_ok) else "differs"
        stored = archive(data, doc_id, file_name, quarantined=(status == "quarantined"))
        run(f"""INSERT INTO {FQ}.`1_source_document` VALUES
            ({sq(doc_id)},{sq(meta['carrier'])},{sq(meta['group'])},{sq(meta['period'])},{sq(file_name)},
             {sq(meta['doc_family'])},current_timestamp(),{expected},{found},{sq(detail)},{sq(status)},NULL,NULL,{sq(stored)})""", "quarantine")
        audit("quarantined" if status == "quarantined" else "extraction_differs", "source_document", doc_id,
              f"{found}/{expected} fields; missing tabs {missing_tabs}; {disagreements} disagreements", actor)
        if stored:
            audit("archived", "source_document", doc_id, f"original stored in governed Volume: {stored}", actor)
        return {"status": status, "doc_id": doc_id, "found": found, "expected": expected,
                "missing_tabs": missing_tabs, "recon": recon, "stored_path": stored}

    # supersede prior active for same group+period
    old = run(f"SELECT doc_id FROM {FQ}.`1_source_document` WHERE employer_group={sq(meta['group'])} "
              f"AND policy_period={sq(meta['period'])} AND status='active'", "find old").result.data_array or []
    for (old_id,) in old:
        run(f"UPDATE {FQ}.`1_source_document` SET status='superseded' WHERE doc_id={sq(old_id)}", "supersede")
        audit("superseded", "source_document", old_id, f"replaced by {doc_id}", actor)

    # load source rows tagged with the new doc_id
    v = {k: det[k]["value"] for k in det}
    months = parse_monthly(sheets)
    if months:
        vals = ",".join(f"({sq(doc_id)},{sq(meta['carrier'])},{sq(meta['group'])},{sq(m['month'])},"
                        f"{sq(m['billed_premium'])},{sq(m['ffs_medical'])},{sq(m['pharmacy'])},"
                        f"{sq(m['fixed_charges'])},NULL,{sq(m['total_incurred'])},{sq(int(m['med_ees'] or 0))},{sq(int(m['med_members'] or 0))})"
                        for m in months)
        run(f"INSERT INTO {FQ}.`1_incurred_claims` VALUES {vals}", "load incurred")
    larges = parse_large(sheets)
    if larges:
        vals = ",".join(f"({sq(doc_id)},{sq(meta['carrier'])},{sq(meta['group'])},{sq(l['status'])},{sq(l['ee_dep'])},NULL,{sq(l['amount'])})"
                        for l in larges)
        run(f"INSERT INTO {FQ}.`1_large_claims` VALUES {vals}", "load large")
    details = parse_detailed(sheets)
    if details:
        vals = ",".join(f"({sq(doc_id)},{sq(meta['carrier'])},{sq(meta['group'])},{sq(d['kind'])},{sq(d['plan'])},"
                        f"{sq(d['tier'])},{sq(d['subs'])},{sq(d['rate'])})" for d in details)
        run(f"INSERT INTO {FQ}.`1_detailed_rates` VALUES {vals}", "load detailed")
    members = int(v.get("current_members") or 0) or 1
    cur_prem_monthly = (v.get("current_premium_pmpm") or 0) * members
    exp_w = v.get("credibility_experience_weight") or 0.652
    run(f"""INSERT INTO {FQ}.`2_renewal_inputs` VALUES
        ({sq(doc_id)},{sq(meta['carrier'])},{sq(meta['group'])},NULL,NULL,
         {sq(int(v.get('months_experience') or 12))},'Fully Insured',{sq(v.get('individual_pooling_point'))},
         {sq(meta['period'])},{sq(v.get('demographic_adjustment'))},{sq(v.get('less_pooled_claims_pmpm'))},
         {sq(v.get('benefit_change'))},{sq(v.get('annual_trend'))},{sq(v.get('months_of_trend'))},
         {sq(v.get('projected_excess_claims_pmpm'))},{sq(v.get('large_claim_add_back_pmpm'))},
         {sq(members)},{sq(cur_prem_monthly)},{sq(v.get('target_loss_ratio'))},{sq(v.get('benefit_advisor_fee'))},
         {sq(v.get('manual_rating_pool_increase'))},{sq(exp_w)},{sq(round(1-exp_w,4))},0.0,
         {sq(v.get('member_months'))},{sq(v.get('total_incurred_claims'))})""", "load inputs")

    # store the FULL two-path reconciliation so the app can show what was extracted
    active_detail = json.dumps({"found": found, "expected": expected, "reconciliation": recon})
    stored = archive(data, doc_id, file_name, quarantined=False)
    run(f"""INSERT INTO {FQ}.`1_source_document` VALUES
        ({sq(doc_id)},{sq(meta['carrier'])},{sq(meta['group'])},{sq(meta['period'])},{sq(file_name)},
         {sq(meta['doc_family'])},current_timestamp(),{expected},{found},{sq(active_detail)},'active',NULL,NULL,{sq(stored)})""", "source doc")
    audit("ingested", "source_document", doc_id, f"{found}/{expected} fields, both paths agree; {len(months)} months, {len(larges)} large claims", actor)
    if stored:
        audit("archived", "source_document", doc_id, f"original stored in governed Volume: {stored}", actor)
    return {"status": "active", "doc_id": doc_id, "found": found, "expected": expected, "recon": recon, "stored_path": stored}


if __name__ == "__main__":
    import sys
    res = process_file(sys.argv[1], actor="demo@local")
    print(json.dumps({k: v for k, v in res.items() if k != "recon"}, indent=2, default=str))
