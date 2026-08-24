"""Generate the synthetic demo carrier files for the ingestion beats.

Produces four workbooks (fictional carrier + employer, invented figures):
  * meridian_harborview_2026H2.xlsx        — hero file, clean, matches template
  * meridian_harborview_2026H2_v2.xlsx     — same group/period, trend changed (supersession)
  * meridian_brokenlayout.xlsx             — malformed (quarantine beat)
  * meridian_harborview_2026H2_month13.xlsx— one fresh claims month (reforecast beat)

Filename convention: <carrier>_<group>_<period>.xlsx
Layout deliberately differs from any mapped reference: different sheet names,
label wording and row order — extraction must work on a layout nobody hand-mapped.
"""
from __future__ import annotations
import os
import openpyxl
from openpyxl.styles import Font, PatternFill

CARRIER = "Meridian Assurance"
CLIENT = "Harborview Logistics"

BASE = dict(
    member_months=5_412, months_experience=12, total_incurred_annual=4_186_930.44,
    demographic_adj=1.0142, pooled_credit_pmpm=158.75, benefit_change=0.9840,
    annual_trend=0.1220, months_of_trend=15, pooling_point=125_000,
    excess_pmpm=133.40, add_back_pmpm=111.25, current_members=462,
    target_lr=0.9180, advisor_fee=0.045, manual_increase=0.1735,
    experience_weight=0.71, current_premium_pmpm=831.40,
)
HDR = Font(bold=True, size=12); LBL = Font(bold=True); FILL = PatternFill("solid", fgColor="E8EEF7")

MONTHS = ["Oct-2025", "Nov-2025", "Dec-2025", "Jan-2026", "Feb-2026", "Mar-2026",
          "Apr-2026", "May-2026", "Jun-2026", "Jul-2026", "Aug-2026", "Sep-2026"]
WEIGHTS = [.081, .078, .088, .095, .086, .084, .082, .079, .083, .081, .080, .083]
MEMBERS = [446, 448, 450, 452, 455, 452, 450, 453, 456, 458, 460, 462]
CLAIMANTS = [
    ("HC-0413", "Active", "EE", "Oncology", 287_412.55), ("HC-0511", "Active", "DEP", "Neonatal", 244_180.10),
    ("HC-0298", "Termed", "EE", "Cardiac", 196_755.80), ("HC-0677", "Active", "DEP", "Transplant", 173_920.35),
    ("HC-0102", "Active", "EE", "Musculoskeletal", 141_265.90), ("HC-0845", "Active", "DEP", "Haemophilia", 128_640.75),
    ("HC-0390", "Active", "EE", "Oncology", 118_405.20), ("HC-0733", "Termed", "DEP", "Sepsis", 96_310.45),
    ("HC-0166", "Active", "EE", "Autoimmune", 84_775.60), ("HC-0904", "Active", "DEP", "Prematurity", 71_240.15),
    ("HC-0455", "Active", "EE", "Diabetes", 63_890.05), ("HC-0588", "Active", "DEP", "Bariatric", 52_115.70),
]


def _derive(p):
    incurred_pmpm = p["total_incurred_annual"] / p["member_months"]
    adjusted = incurred_pmpm * p["demographic_adj"]
    exp_claim_cost = adjusted - p["pooled_credit_pmpm"]
    eff_trend = (1 + p["annual_trend"]) ** (p["months_of_trend"] / 12) - 1
    projected = exp_claim_cost * p["benefit_change"] * (1 + eff_trend) + p["excess_pmpm"] + p["add_back_pmpm"]
    exp_premium = projected / p["target_lr"] * (1 + p["advisor_fee"])
    exp_increase = exp_premium / p["current_premium_pmpm"] - 1
    blended = p["experience_weight"] * exp_increase + (1 - p["experience_weight"]) * p["manual_increase"]
    return dict(incurred_pmpm=incurred_pmpm, adjusted=adjusted, exp_claim_cost=exp_claim_cost,
                eff_trend=eff_trend, projected=projected, exp_premium=exp_premium,
                exp_increase=exp_increase, blended=blended)


def _claims_sheet(wb, months, weights, members, total):
    ws = wb.create_sheet("Claims Experience")
    ws.append(["Month", "Billed Premium", "Medical Claims", "Rx Claims", "Admin & Fixed",
               "Total Incurred", "Subscribers", "Covered Members"])
    for c in ws[1]:
        c.font = LBL; c.fill = FILL
    for i, m in enumerate(months):
        tot = total * weights[i]
        med, rx = tot * 0.615, tot * 0.335
        ws.append([m, round(tot / 1.09, 2), round(med, 2), round(rx, 2), round(tot - med - rx, 2),
                   round(tot, 2), round(members[i] * 0.61), members[i]])
    return ws


# --- synthetic medical plans (generic names — no real product brands) ---
PLANS = [
    # (plan code, product, network, [current tier rates], subs per tier)
    ("PPO",     "Preferred PPO",        "National PPO", [860.00, 2050.00, 1585.00, 2690.00], [22, 4, 3, 7]),
    ("PPO-IN",  "PPO In-Network",       "National PPO", [790.00, 1880.00, 1455.00, 2470.00], [130, 5, 6, 3]),
    ("HDHP",    "HDHP with HSA",        "National PPO", [690.00, 1645.00, 1270.00, 2155.00], [58, 2, 4, 6]),
]
TIERS = ["Employee", "Emp + Spouse", "Emp + Child(ren)", "Emp + Family"]
RENEWAL_FACTOR = 1.269  # illustrative per-contract renewal increase (echoes the blended action)
BENEFIT_LINES = [  # (label, PPO value, PPO-IN value, HDHP value)
    ("Deductible (Single/Family)", "$2,500 / $5,000", "$1,500 / $3,000", "$3,000 / $6,000"),
    ("Out-of-Pocket (Single/Family)", "$7,500 / $15,000", "$7,500 / $15,000", "$6,500 / $13,000"),
    ("Coinsurance", "80%", "80%", "70%"),
    ("Physician Services - PCP", "$30 copay", "$40 copay", "70% ^"),
    ("Physician Services - SPC", "$60 copay", "$70 copay", "70% ^"),
    ("Inpatient Services", "80% ^", "80% ^", "70% ^"),
    ("Outpatient Services", "80% ^", "80% ^", "70% ^"),
    ("Emergency Room", "80% ^", "80% ^", "70% ^"),
    ("Urgent Care", "$75 copay", "$75 copay", "70% ^"),
    ("Virtual Visit - Urgent Care", "100%", "100%", "100% ^"),
    ("Lab Services", "100%", "100%", "70% ^"),
    ("Adv. Radiology - Outpatient", "80% ^", "80% ^", "70% ^"),
    ("Outpatient PT / OT / Speech", "Same as Spc. OV", "Same as Spc. OV", "Same as Spc. OV"),
    ("Chiropractic Care", "Plan Coins", "Plan Coins", "Plan Coins"),
]


def _benefits_sheet(wb):
    ws = wb.create_sheet("Benefits Overview")
    ws["A1"] = CLIENT; ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Benefits & Total Rate Overview: Fully Insured"; ws["A2"].font = LBL
    ws["A3"] = "2026-H2 Renewal"
    hdr = ["Plan"] + [pl[0] for pl in PLANS]
    ws.append([]); ws.append(hdr)
    for c in ws[5]:
        c.font = LBL; c.fill = FILL
    ws.append(["Product"] + [pl[1] for pl in PLANS])
    ws.append(["Network"] + [pl[2] for pl in PLANS])
    ws.append(["In-Network"]); ws.cell(ws.max_row, 1).font = LBL
    for line in BENEFIT_LINES:
        ws.append(list(line))
    ws.append(["Pharmacy"]); ws.cell(ws.max_row, 1).font = LBL
    ws.append(["Pharmacy Network", "Retail + 90-day", "Retail + 90-day", "Retail + 90-day"])
    ws.append(["Formulary", "Advantage", "Advantage", "Advantage"])
    for col, wdt in zip("ABCD", (34, 26, 26, 26)):
        ws.column_dimensions[col].width = wdt
    return ws


def _detailed_rates_sheet(wb):
    ws = wb.create_sheet("Detailed Rates")
    ws["A1"] = CLIENT; ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Detailed Rates : Fully Insured"; ws["A2"].font = LBL
    ws["A3"] = "2026-H2 Renewal"
    for kind, label, factor in [("current", "CURRENT RATES", 1.0), ("renewal", "RENEWAL RATES", RENEWAL_FACTOR)]:
        ws.append([]); ws.append([label]); ws.cell(ws.max_row, 1).font = HDR
        for code, product, network, rates, subs in PLANS:
            ws.append([f"Plan: {code}", f"Product:   {product}"]); ws.cell(ws.max_row, 1).font = LBL
            ws.append([f"Network:   {network}"])
            ws.append(["Subs", "Total"]);
            for c in ws[ws.max_row]:
                c.font = LBL
            total_subs, total_cost = 0, 0.0
            for tier, rate, n in zip(TIERS, rates, subs):
                r = round(rate * factor, 2)
                ws.append([tier, n, r]); total_subs += n; total_cost += n * r
            ws.append(["Monthly Plan Cost", total_subs, round(total_cost, 2)]); ws.cell(ws.max_row, 1).font = LBL
    for col, wdt in zip("ABC", (22, 12, 14)):
        ws.column_dimensions[col].width = wdt
    return ws


def build(path, p, *, broken=False):
    d = _derive(p)
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Renewal Summary"
    ws["A1"] = CLIENT; ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"{CARRIER} — Medical Renewal Illustration"
    ws["A3"] = "Funding Arrangement: Fully Insured"
    ws["A4"] = "Policy Period Commencing: 2026-H2"
    ws["A6"] = "SYNTHETIC EXAMPLE — fictional carrier and employer, invented figures."
    ws["A6"].font = Font(italic=True, size=9); ws.column_dimensions["A"].width = 52

    _claims_sheet(wb, MONTHS, WEIGHTS, MEMBERS, p["total_incurred_annual"])

    ws = wb.create_sheet("High Cost Claimants")
    ws.append(["Claimant Ref", "Status", "Relationship", "Primary Diagnosis", "Paid Amount"])
    for c in ws[1]:
        c.font = LBL; c.fill = FILL
    for c in CLAIMANTS:
        ws.append(list(c))
    r = ws.max_row + 2
    ws.cell(r, 1, "Individual Pooling Point").font = LBL
    ws.cell(r, 5, p["pooling_point"])

    ws = wb.create_sheet("Rate Development")
    ws["A1"] = f"{CLIENT} — Rate Development"; ws["A1"].font = HDR
    # Broken: rename a key label and drop Member Months so extraction misses fields.
    trend_label = "Trend Rate (annual)" if broken else "Annual Trend"
    rows = [
        ("Total Months of Experience", p["months_experience"], None),
        (None if broken else "Member Months", p["member_months"], None),
        ("Individual Pooling Point", p["pooling_point"], None),
        ("Current members", p["current_members"], None),
        ("Total Incurred Claims", round(d["incurred_pmpm"], 6), round(p["total_incurred_annual"], 2)),
        ("Demographic Adjustment", p["demographic_adj"], None),
        ("Adjusted Incurred Claims", round(d["adjusted"], 6), None),
        ("Less Pooled Claims", p["pooled_credit_pmpm"], None),
        ("Experience Claim Cost", round(d["exp_claim_cost"], 6), None),
        ("Benefit Change", p["benefit_change"], None),
        (trend_label, p["annual_trend"], None),
        ("Months of Trend", p["months_of_trend"], None),
        ("Effective Trend", round(d["eff_trend"], 8), None),
        (f"Projected claims in excess of ${p['pooling_point']:,}", p["excess_pmpm"], None),
        ("Large Claim Add Back", p["add_back_pmpm"], None),
        ("Projected Medical Costs", round(d["projected"], 6), None),
        ("Target Loss Ratio", p["target_lr"], None),
        ("Benefit Advisor Fees", p["advisor_fee"], None),
        ("Experience Based Premium & Fees", round(d["exp_premium"], 6), None),
        ("Current Premium & Fees", p["current_premium_pmpm"], None),
        ("Experience Based Rate Increase", round(d["exp_increase"], 8), None),
        ("Manual Rating Pool Increase", p["manual_increase"], None),
    ]
    row = 4 if not broken else 6  # broken: shift start row too
    for label, pmpm, annual in rows:
        if label is None:
            row += 1; continue
        ws.cell(row, 1, label); ws.cell(row, 2, pmpm)
        if annual is not None:
            ws.cell(row, 3, annual)
        row += 1
    row += 1
    ws.cell(row, 1, "Renewal Action Basis").font = LBL; ws.cell(row, 2, "Credibility Weighted")
    ws.cell(row + 1, 1, "Experience Based Rate Increase"); ws.cell(row + 1, 2, p["experience_weight"]); ws.cell(row + 1, 3, round(d["exp_increase"], 8))
    ws.cell(row + 2, 1, "Manual Rating Pool Increase"); ws.cell(row + 2, 2, round(1 - p["experience_weight"], 4)); ws.cell(row + 2, 3, p["manual_increase"])
    ws.cell(row + 4, 1, "Blended Renewal Rate Action").font = LBL; ws.cell(row + 4, 3, round(d["blended"], 8)).font = LBL
    ws.cell(row + 5, 1, "Adjustment"); ws.cell(row + 5, 3, 0)
    ws.cell(row + 6, 1, "Quoted Change in Billed Amount").font = LBL; ws.cell(row + 6, 3, round(d["blended"], 8)).font = LBL
    ws.column_dimensions["A"].width = 44; ws.column_dimensions["B"].width = 20; ws.column_dimensions["C"].width = 20
    _benefits_sheet(wb)
    _detailed_rates_sheet(wb)
    wb.save(path)
    return d


def build_month13(path, p):
    """One fresh claims month for the same group — the reforecast document."""
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    _claims_sheet(wb, ["Oct-2026"], [1.0], [464], p["total_incurred_annual"] * 0.079)
    wb.save(path)


if __name__ == "__main__":
    # output dir: HB_DEMO_DIR (writable /tmp on a Job) else next to this file
    here = os.getenv("HB_DEMO_DIR") or os.path.dirname(os.path.abspath(__file__))
    os.makedirs(here, exist_ok=True)
    hero = dict(BASE)
    d = build(os.path.join(here, "meridian_harborview_2026H2.xlsx"), hero)
    print(f"hero blended action: {d['blended']*100:.4f}%")
    v2 = dict(BASE, annual_trend=0.1050)
    build(os.path.join(here, "meridian_harborview_2026H2_v2.xlsx"), v2)
    build(os.path.join(here, "meridian_brokenlayout.xlsx"), dict(BASE), broken=True)
    build_month13(os.path.join(here, "meridian_harborview_2026H2_month13.xlsx"), hero)
    print("wrote 4 demo files.")
